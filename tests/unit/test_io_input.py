"""Unit tests for io_input.py — covers the Anna 2026-05-13 input model changes:
email, optional geburtsjahr, nachbelastungslaktat, hh:mm validation, Anfangsbelastung."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from ld import io_input
from ld.errors import LDInputError


FIXTURE = Path(__file__).parent.parent / "protocols" / "fixtures" / "lauf" / "rainier.xlsx"


def test_email_optional_default_none():
    """Existing fixtures have an empty Email cell → parsed as None."""
    run = io_input.parse_input(FIXTURE)
    assert run.athlete.email is None


def test_geburtsjahr_optional_when_blank(tmp_path):
    """A blank Geburtsjahr cell must parse cleanly and yield alter=None."""
    src = openpyxl.load_workbook(FIXTURE)
    for row in src["Athlet"].iter_rows():
        if row[0].value == "Geburtsjahr":
            row[1].value = None  # clear it
            break
    out = tmp_path / "blank_gj.xlsx"
    src.save(out)
    run = io_input.parse_input(out)
    assert run.athlete.geburtsjahr is None
    assert run.athlete.alter is None


def test_email_parsed_when_present(tmp_path):
    src = openpyxl.load_workbook(FIXTURE)
    for row in src["Athlet"].iter_rows():
        if row[0].value == "Email":
            row[1].value = "rainier@example.com"
            break
    out = tmp_path / "with_email.xlsx"
    src.save(out)
    run = io_input.parse_input(out)
    assert run.athlete.email == "rainier@example.com"


def test_nachbelastungslaktat_optional():
    """Blank Nachbelastungslaktat rows → None on the Testprotokoll dataclass."""
    run = io_input.parse_input(FIXTURE)
    assert run.testprotokoll.nachbelastungslaktat_3min_mmol is None
    assert run.testprotokoll.nachbelastungslaktat_5min_mmol is None


def test_nachbelastungslaktat_parsed(tmp_path):
    src = openpyxl.load_workbook(FIXTURE)
    for row in src["Testprotokoll"].iter_rows():
        if row[0].value == "Nachbelastungslaktat 3min (mmol/l)":
            row[1].value = 6.2
        if row[0].value == "Nachbelastungslaktat 5min (mmol/l)":
            row[1].value = 4.1
    out = tmp_path / "nach.xlsx"
    src.save(out)
    run = io_input.parse_input(out)
    assert run.testprotokoll.nachbelastungslaktat_3min_mmol == 6.2
    assert run.testprotokoll.nachbelastungslaktat_5min_mmol == 4.1


def test_anfangsbelastung_renamed_from_intensitaet():
    """The Testprotokoll-dataclass field is named anfangsbelastung."""
    run = io_input.parse_input(FIXTURE)
    assert run.testprotokoll.anfangsbelastung == 7.0
    assert not hasattr(run.testprotokoll, "anfangsintensitaet")


def test_uhrzeit_rejects_non_hhmm(tmp_path):
    src = openpyxl.load_workbook(FIXTURE)
    for row in src["Testprotokoll"].iter_rows():
        if row[0].value == "Uhrzeit":
            row[1].value = "2pm"  # invalid format
            break
    out = tmp_path / "bad_uhrzeit.xlsx"
    src.save(out)
    with pytest.raises(LDInputError, match="hh:mm"):
        io_input.parse_input(out)
