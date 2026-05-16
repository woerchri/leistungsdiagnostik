"""RPE 0-10 (Borg CR10) range validation — Anna 2026-05-13 Round 2.

Old Borg 6-20 values would silently corrupt zone metadata (which is set up for
0-10). The parser rejects out-of-range RPE explicitly so a stale template is
caught at parse time rather than at zone-rendering time.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from ld import io_input
from ld.errors import LDInputError


FIXTURE = Path(__file__).parent.parent / "protocols" / "fixtures" / "lauf" / "rainier.xlsx"


def test_rpe_within_range_accepted():
    """The cleaned fixture should parse without error."""
    run = io_input.parse_input(FIXTURE)
    assert all(s.rpe is None or 0 <= s.rpe <= 10 for s in run.steps)


def test_rpe_borg_value_rejected(tmp_path):
    """A Borg 6-20 value (e.g. 18) must be rejected with a clear message."""
    src = openpyxl.load_workbook(FIXTURE)
    ws = src["Testdaten"]
    ws.cell(row=2, column=5, value=18)  # old Borg value
    out = tmp_path / "borg_rpe.xlsx"
    src.save(out)
    with pytest.raises(LDInputError, match="CR10"):
        io_input.parse_input(out)


def test_rpe_negative_rejected(tmp_path):
    src = openpyxl.load_workbook(FIXTURE)
    ws = src["Testdaten"]
    ws.cell(row=2, column=5, value=-1)
    out = tmp_path / "neg_rpe.xlsx"
    src.save(out)
    with pytest.raises(LDInputError, match="CR10"):
        io_input.parse_input(out)


def test_rpe_missing_allowed(tmp_path):
    """RPE is optional per-step — empty cells stay None."""
    src = openpyxl.load_workbook(FIXTURE)
    ws = src["Testdaten"]
    # Direct attribute assignment; `ws.cell(..., value=None)` does NOT overwrite
    # in openpyxl (None is treated as "leave existing").
    ws.cell(row=2, column=5).value = None
    out = tmp_path / "blank_rpe.xlsx"
    src.save(out)
    run = io_input.parse_input(out)
    assert run.steps[0].rpe is None
