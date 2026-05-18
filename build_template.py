"""Build templates/input_template.xlsx pre-filled with Rainier's data."""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("templates/input_template.xlsx")
OUT.parent.mkdir(exist_ok=True)

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
NOTE_FILL   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
KEY_FILL    = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
KEY_FONT    = Font(bold=True, name="Arial")
BODY_FONT   = Font(name="Arial")

thin = Side(style="thin", color="CCCCCC")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal="center")

def kv(ws, row, key, val, note=None):
    kc = ws.cell(row=row, column=1, value=key)
    kc.font = KEY_FONT; kc.fill = KEY_FILL
    vc = ws.cell(row=row, column=2, value=val)
    vc.font = BODY_FONT
    if note:
        nc = ws.cell(row=row, column=3, value=note)
        nc.font = Font(name="Arial", italic=True, color="808080")

# ── Athlet ──────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Athlet"
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 45

hdr(ws, 1, 1, "Feld"); hdr(ws, 1, 2, "Wert"); hdr(ws, 1, 3, "Hinweis")

# Anna 2026-05-13 Round 1: "Sportart aus Athletendaten entfernen und bei
# Testprotokoll anzeigen" — moved to Testprotokoll sheet below.
rows = [
    ("Vorname",            "Rainier",        ""),
    ("Name",               "Matzinger",      ""),
    ("Email",              "rainier@example.com", "Optional"),
    ("Geburtsjahr",        1968,             "Vierstellig, z.B. 1985 — optional (kann leer bleiben)"),
    ("Geschlecht",         "m",              "m | w | d (optional)"),
    ("Gewicht (kg)",       71.5,             ""),
    ("Größe (m)",          1.79,             "In Metern, z.B. 1.79"),
    ("Trainingsziel",      "Marathonzeit verbessern", ""),
    ("Wettkampfziel",      "Achenseelauf 2024",        ""),
    ("Trainingsumfang/Woche", "ca. 8h",      ""),
    ("Leistungsniveau",    "Hobbyläufer",    ""),
]
for i, (k, v_, n) in enumerate(rows, start=2):
    kv(ws, i, k, v_, n)

# ── Testprotokoll ────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Testprotokoll")
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 45

hdr(ws2, 1, 1, "Feld"); hdr(ws2, 1, 2, "Wert"); hdr(ws2, 1, 3, "Hinweis")

import datetime
proto_rows = [
    ("Sportart",                           "Lauf",                        "Erlaubt: lauf | rad | triathlon-rad | triathlon-lauf | unspezifisch"),
    ("Testdatum",                          "23.05.2024",                  "Format: TT.MM.JJJJ"),
    ("Uhrzeit",                            "14:00",                       "Format: hh:mm"),
    ("Durchführungsort",                   "Laufcamp Achensee",           ""),
    ("Testleiter",                         "Anna-Maria Wörndle",          ""),
    ("Gerät",                              "Laufband Atoll Achensee",     ""),
    ("Anfangsbelastung",                   7.0,                           "km/h (Lauf) | W (Rad) | Stufe (Unspez)"),
    ("Stufeninkrement",                    1.0,                           "km/h | W | Stufen"),
    ("Stufendauer (min)",                  4.0,                           "Minuten"),
    ("Stufenlänge (m)",                    None,                          "Optional, nur Lauf (z.B. 1600)"),
    # Round 4 (Anna 2026-05-18): Steigung hat ein eigenes Feld unten —
    # daher hier keinen "1% Steigung"-Beispieltext mehr (vermeidet doppelte
    # Eingabe und macht klar, dass Steigung strukturiert erfasst wird).
    ("Besonderheiten",                     "",                            "Freitext (z.B. Witterung, Anreise)"),
    ("Letzte Stufe vollständig absolviert","Nein",                        "JA oder NEIN"),
    ("Dauer letzte Stufe (min)",           1.75,                          "Nur ausfüllen wenn NEIN oben"),
    ("Ausbelastung",                       "Ja",                          "JA oder NEIN"),
    ("Nachbelastungslaktat 3min (mmol/l)", None,                          "Optional — Laktatwert 3 min nach Belastung"),
    ("Nachbelastungslaktat 5min (mmol/l)", None,                          "Optional — Laktatwert 5 min nach Belastung"),
    # Round 3 (Anna 2026-05-17): Ruhelaktat + Steigung als eigene Felder.
    ("Ruhelaktat (mmol/l)",                1.0,                           "Optional — Laktat vor Testbeginn"),
    ("Steigung (%)",                       1.0,                           "Optional — Laufband-/Strecken-Steigung in Prozent"),
]
for i, (k, v_, n) in enumerate(proto_rows, start=2):
    kv(ws2, i, k, v_, n)

# ── Testdaten ────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Testdaten")
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 8

headers = ["Stufe", "Intensität", "Herzfrequenz", "Laktat", "RPE"]
for col, h in enumerate(headers, start=1):
    hdr(ws3, 1, col, h)

# RPE note in column G (outside the data range, visible to user).
# Anna 2026-05-13 Round 2: full migration to Borg CR10 (0-10) — old 6-20 is rejected.
note_cell = ws3.cell(row=1, column=7,
    value="RPE-Skala: 0-10 (Borg CR10). 0 = keine Anstrengung, 10 = maximale Anstrengung")
note_cell.fill = NOTE_FILL
note_cell.font = Font(name="Arial", italic=True, color="7F6000")

# Rainier's data — 6 measured steps (step 6 is incomplete; aliquot vmax).
# RPE values on CR10 0-10 (Borg→CR10 conversion: 9→2, 14→5, 16→7, 17→8, 18→9).
data = [
    (1, 7.0,  133, 1.9, 2),
    (2, 8.0,  141, 2.8, 2),
    (3, 9.0,  150, 3.0, 5),
    (4, 10.0, 159, 3.8, 7),
    (5, 11.0, 167, 5.2, 8),
    (6, 12.0, 173, 7.9, 9),
]
for i, row_data in enumerate(data, start=2):
    for col, val in enumerate(row_data, start=1):
        c = ws3.cell(row=i, column=col, value=val)
        c.font = BODY_FONT

# ── Coaching ─────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("Coaching")
ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 50

hdr(ws4, 1, 1, "Feld"); hdr(ws4, 1, 2, "Wert")

coaching_rows = [
    ("Verletzungen",       "keine"),
    ("Aktuelle Probleme",  ""),
    ("Stärken",            "Grundlagenausdauer"),
    ("Schwächen",          "Tempohärte"),
    ("Geplante Wettkämpfe","Achenseelauf Juni 2024"),
    ("Trainernotizen",     "Erste Diagnostik der Saison"),
]
for i, (k, v_) in enumerate(coaching_rows, start=2):
    kc = ws4.cell(row=i, column=1, value=k)
    kc.font = KEY_FONT; kc.fill = KEY_FILL
    vc = ws4.cell(row=i, column=2, value=v_)
    vc.font = BODY_FONT

wb.save(OUT)
print(f"Template gespeichert: {OUT}")
