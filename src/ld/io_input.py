from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path

import openpyxl

from ld.errors import LDInputError
from ld.types import Athlete, Coaching, TestRun, TestStep, Testprotokoll


def parse_input(path: Path) -> TestRun:
    if not path.exists():
        raise LDInputError(f"Eingabedatei nicht gefunden: {path}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise LDInputError(f"Eingabedatei konnte nicht geöffnet werden ({path.name}): {e}") from e

    return TestRun(
        athlete=_parse_athlete(wb),
        testprotokoll=_parse_testprotokoll(wb),
        steps=_parse_steps(wb),
        coaching=_parse_coaching(wb),
    )


def _require_sheet(wb, name: str):
    if name not in wb.sheetnames:
        raise LDInputError(
            f"Arbeitsblatt '{name}' fehlt in der Eingabedatei. "
            f"Bitte mit templates/input_template.xlsx als Basis arbeiten."
        )
    return wb[name]


def _kv(ws, key: str) -> str | None:
    """Read column-A keyed pairs (key in col A, value in col B). Returns string or None."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row and row[0] == key:
            return row[1]
    return None


def _required_kv(ws, key: str) -> str:
    value = _kv(ws, key)
    if value is None or value == "":
        raise LDInputError(f"Pflichtfeld '{key}' auf Blatt '{ws.title}' ist leer.")
    return str(value).strip()


def _parse_athlete(wb) -> Athlete:
    ws = _require_sheet(wb, "Athlet")
    # Sportart lives on the Testprotokoll sheet (Anna 2026-05-13 Round 1).
    # Legacy templates may still have it on Athlet — accept either to keep old
    # fixtures parseable, but prefer the Testprotokoll location.
    proto_ws = _require_sheet(wb, "Testprotokoll")
    sportart_raw_value = _kv(proto_ws, "Sportart") or _kv(ws, "Sportart")
    if not sportart_raw_value:
        raise LDInputError(
            "Pflichtfeld 'Sportart' fehlt — bitte im Testprotokoll-Blatt ergänzen."
        )
    sportart_raw = str(sportart_raw_value).lower().strip()
    SUPPORTED = {"lauf", "rad", "triathlon-rad", "triathlon-lauf", "unspezifisch"}
    if sportart_raw not in SUPPORTED:
        raise LDInputError(
            f"Sportart '{sportart_raw}' nicht unterstützt. Erlaubt: {sorted(SUPPORTED)}."
        )
    geburtsjahr_raw = _kv(ws, "Geburtsjahr")  # Optional per Anna 2026-05-13 feedback
    geburtsjahr: int | None
    if geburtsjahr_raw is None or geburtsjahr_raw == "":
        geburtsjahr = None
    else:
        try:
            geburtsjahr = int(geburtsjahr_raw)
        except (ValueError, TypeError) as e:
            raise LDInputError(
                f"Geburtsjahr muss vierstellig sein oder leer bleiben, nicht '{geburtsjahr_raw}'."
            ) from e
    email_raw = _kv(ws, "Email") or _kv(ws, "E-Mail")  # Both spellings accepted
    try:
        return Athlete(
            sportart=sportart_raw,
            vorname=_required_kv(ws, "Vorname"),
            name=_required_kv(ws, "Name"),
            geburtsjahr=geburtsjahr,
            geschlecht=(_kv(ws, "Geschlecht") or None),
            gewicht_kg=float(_required_kv(ws, "Gewicht (kg)")),
            groesse_m=float(_required_kv(ws, "Größe (m)")),
            trainingsziel=str(_kv(ws, "Trainingsziel") or ""),
            wettkampfziel=str(_kv(ws, "Wettkampfziel") or ""),
            trainingsumfang_wo=str(_kv(ws, "Trainingsumfang/Woche") or ""),
            leistungsniveau=str(_kv(ws, "Leistungsniveau") or ""),
            email=(str(email_raw).strip() if email_raw else None),
        )
    except (ValueError, TypeError) as e:
        raise LDInputError(f"Athletendaten konnten nicht gelesen werden: {e}") from e


def _parse_testprotokoll(wb) -> Testprotokoll:
    ws = _require_sheet(wb, "Testprotokoll")
    raw_date = _kv(ws, "Testdatum")
    if isinstance(raw_date, datetime):
        test_date = raw_date.date()
    elif isinstance(raw_date, date):
        test_date = raw_date
    elif isinstance(raw_date, str):
        test_date = datetime.strptime(raw_date.strip(), "%d.%m.%Y").date()
    else:
        raise LDInputError(
            "Testdatum auf Blatt 'Testprotokoll' fehlt oder hat falsches Format (erwartet TT.MM.JJJJ)."
        )

    # Uhrzeit must be hh:mm (Anna 2026-05-13 feedback)
    uhrzeit_raw = _kv(ws, "Uhrzeit") or "00:00"
    if isinstance(uhrzeit_raw, time):
        uhrzeit = uhrzeit_raw.strftime("%H:%M")
    else:
        uhrzeit = str(uhrzeit_raw).strip()
        if not _is_hhmm(uhrzeit):
            raise LDInputError(
                f"Uhrzeit muss im Format hh:mm sein, nicht '{uhrzeit}'."
            )

    def yesno(key: str) -> bool:
        raw = _required_kv(ws, key).lower().strip()
        if raw in {"ja", "yes", "true", "1"}:
            return True
        if raw in {"nein", "no", "false", "0"}:
            return False
        raise LDInputError(
            f"Feld '{key}' auf 'Testprotokoll' erwartet JA oder NEIN, nicht '{raw}'."
        )

    letzte_voll = yesno("Letzte Stufe vollständig absolviert")
    dauer_letzte = None
    if not letzte_voll:
        raw = _required_kv(ws, "Dauer letzte Stufe (min)")
        try:
            dauer_letzte = float(raw)
        except ValueError as e:
            raise LDInputError(
                f"Dauer letzte Stufe muss eine Zahl in Minuten sein, nicht '{raw}'."
            ) from e

    stufenlaenge_raw = _kv(ws, "Stufenlänge (m)")

    # Anfangsbelastung (renamed from Anfangsintensität per Anna 2026-05-13).
    # The new key is preferred; the old key is no longer accepted — input templates
    # are regenerated alongside this change.
    anfangsbelastung_raw = _required_kv(ws, "Anfangsbelastung")

    # Optional post-exertion lactate (Anna 2026-05-13 — recovery indicator).
    nachbel_3 = _float_or_none(_kv(ws, "Nachbelastungslaktat 3min (mmol/l)"))
    nachbel_5 = _float_or_none(_kv(ws, "Nachbelastungslaktat 5min (mmol/l)"))

    return Testprotokoll(
        testdatum=test_date,
        uhrzeit=uhrzeit,
        durchfuehrungsort=_required_kv(ws, "Durchführungsort"),
        testleiter=_required_kv(ws, "Testleiter"),
        geraet=_required_kv(ws, "Gerät"),
        anfangsbelastung=float(anfangsbelastung_raw),
        stufeninkrement=float(_required_kv(ws, "Stufeninkrement")),
        stufendauer_min=float(_required_kv(ws, "Stufendauer (min)")),
        stufenlaenge_m=(int(stufenlaenge_raw) if stufenlaenge_raw else None),
        besonderheiten=str(_kv(ws, "Besonderheiten") or ""),
        letzte_stufe_vollstaendig=letzte_voll,
        dauer_letzte_stufe_min=dauer_letzte,
        ausbelastung=yesno("Ausbelastung"),
        nachbelastungslaktat_3min_mmol=nachbel_3,
        nachbelastungslaktat_5min_mmol=nachbel_5,
    )


_HHMM_RE = re.compile(r"^([01]?\d|2[0-3]):[0-5]\d$")


def _is_hhmm(s: str) -> bool:
    return bool(_HHMM_RE.match(s))


def _parse_steps(wb) -> tuple[TestStep, ...]:
    ws = _require_sheet(wb, "Testdaten")
    expected_headers = ["Stufe", "Intensität", "Herzfrequenz", "Laktat", "RPE"]
    actual = [
        str(ws.cell(row=1, column=c).value).strip()
        if ws.cell(row=1, column=c).value else ""
        for c in range(1, 6)
    ]
    if actual != expected_headers:
        raise LDInputError(
            f"Spaltenüberschriften auf 'Testdaten' falsch. "
            f"Erwartet: {expected_headers}, gefunden: {actual}."
        )

    steps: list[TestStep] = []
    for row_idx in range(2, ws.max_row + 1):
        stufe_val = ws.cell(row=row_idx, column=1).value
        if stufe_val is None or stufe_val == "":
            break
        try:
            rpe_val = _int_or_none(ws.cell(row=row_idx, column=5).value)
            # RPE 0-10 (Borg CR10) — Anna 2026-05-13 Round 2.
            # Old Borg 6-20 scale values would silently corrupt zone metadata;
            # reject explicitly so the user notices the migration.
            if rpe_val is not None and not (0 <= rpe_val <= 10):
                raise LDInputError(
                    f"Zeile {row_idx} auf 'Testdaten': RPE {rpe_val} liegt "
                    f"außerhalb der CR10-Skala (0-10). Bitte alte Borg-6-20-"
                    f"Werte auf 0-10 umrechnen."
                )
            steps.append(TestStep(
                stufe=int(stufe_val),
                intensitaet=float(ws.cell(row=row_idx, column=2).value),
                herzfrequenz_bpm=_int_or_none(ws.cell(row=row_idx, column=3).value),
                laktat_mmol=_float_or_none(ws.cell(row=row_idx, column=4).value),
                rpe=rpe_val,
            ))
        except LDInputError:
            raise
        except (ValueError, TypeError) as e:
            raise LDInputError(
                f"Zeile {row_idx} auf 'Testdaten' enthält ungültige Werte: {e}"
            ) from e

    if len(steps) < 4:
        raise LDInputError(
            f"Mindestens 4 Stufen mit Laktatwerten erforderlich (kubische Anpassung). "
            f"Gefunden: {len(steps)}."
        )
    return tuple(steps)


def _parse_coaching(wb) -> Coaching:
    ws = _require_sheet(wb, "Coaching")
    return Coaching(
        verletzungen=str(_kv(ws, "Verletzungen") or ""),
        aktuelle_probleme=str(_kv(ws, "Aktuelle Probleme") or ""),
        staerken=str(_kv(ws, "Stärken") or ""),
        schwaechen=str(_kv(ws, "Schwächen") or ""),
        geplante_wettkaempfe=str(_kv(ws, "Geplante Wettkämpfe") or ""),
        trainernotizen=str(_kv(ws, "Trainernotizen") or ""),
    )


def _int_or_none(v):
    if v is None or v == "":
        return None
    return int(v)


def _float_or_none(v):
    if v is None or v == "":
        return None
    return float(v)
